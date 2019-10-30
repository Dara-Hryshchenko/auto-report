import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.styles import NamedStyle
import numpy as np
import pandas as pd
from dateutil import parser
import datetime
import win32com.client as win32
try:
    from Tkinter import *
except ImportError:
    from tkinter import *
from tkinter import messagebox

# Constants
# - defining key-word combinations to indentify the files:
choco_daily_file_key = 'choco daily shipments'
pf_daily_file_key = 'pf daily shipments'
operations_SIT_file_key = 'operations'
outb_actual_file_key = 'plan'
inb_actual_file_key = 'inbound _outbound'
stock_monitor_file_key = 'stock'
#- defining sheet names:
choco_daily_sales_sheet_key = "Sales Forecast"
pf_daily_sales_sheet_key = "Sales Forecast"
choco_daily_ship_sheet_key = "Shipments"
pf_daily_ship_sheet_key = "Shipments"
inb_actual_sheet_key = "Inbound"
operations_SIT_sheet_key = "кол-во пал"

## Define functions

# if offset_val is positive then next date is returned
def date_offset(date_val,offset_val):
    date_val += datetime.timedelta(days=offset_val)
    return(date_val)

# convert xls file to xlsx file
def convert_xls_to_xlsx(fname,input_folder):
    try:
        #should br defined full path to work, else the error rises
        fname = input_folder +"\\"+fname
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        #excel = win32.DispatchEx("Excel.Application")
        wb = excel.Workbooks.Open(fname)
        wb.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
        wb.Close()
        #wb.Close(True)                               #FileFormat = 56 is for .xls extension
        # if use next command to quite it will also close all other excel files, opened by user
        excel.Application.Quit()
    except ValueError:
        messagebox.showinfo("Warning", "Please manually change xls files to xlsx and DELETE xls! ")

# to read excel to df without hidden rows and cols
def read_excel_wo_hidden(loc, sheet_name=0, encoding = "latin-1", cols_wo_hidden = True, rows_wo_hidden = True):
    wb = load_workbook(loc)
    if sheet_name == 0:
        sheet_name = wb.sheetnames[0]
    ws = wb.get_sheet_by_name(sheet_name)
    df = pd.read_excel(loc, sheet_name = sheet_name, encoding = encoding)
    if cols_wo_hidden == True:
        hidden_cols = []
        for colLetter,colDimension in ws.column_dimensions.items():
            if colDimension.hidden == True:
                hidden_cols.append(colLetter)
        unhidden_cols = [x for x in df.columns if x not in hidden_cols]
    if rows_wo_hidden == True:
        hidden_rows = []
        for rowNumber,rowDimension in ws.row_dimensions.items():
            if rowDimension.hidden == True:
                hidden_rows.append(rowNumber-2)
        unhidden_rows = [x for x in df.index if x not in hidden_rows]
    if cols_wo_hidden == True and rows_wo_hidden == True:
        df = df.loc[unhidden_rows, unhidden_cols]
    elif cols_wo_hidden == True and rows_wo_hidden == False:
        df = df.loc[:, unhidden_cols]
    elif cols_wo_hidden == False and rows_wo_hidden == True:
        df = df.loc[unhidden_rows, :]
    elif cols_wo_hidden == False and rows_wo_hidden == False:
        df = df.loc[:, :]
    # check for null values
    for col in df.select_dtypes("object"):
        df[col] = np.where(df[col] == '', np.nan, df[col])
        df[col] = np.where(df[col] == ' ', np.nan, df[col])
        df[col] = np.where(df[col] == '  ', np.nan, df[col])
        df[col] = np.where(df[col] == '   ', np.nan, df[col])
    return df

def multiindex_pivot(df, columns=None, values=None):
    names = list(df.index.names)
    df = df.reset_index()
    list_index = df[names].values
    tuples_index = [tuple(i) for i in list_index] # hashable
    df = df.assign(tuples_index=tuples_index)
    df = df.pivot(index="tuples_index", columns=columns, values=values)
    tuples_index = df.index
    index = pd.MultiIndex.from_tuples(tuples_index, names=names)
    df.index = index
    return df


def files_check_and_list(input_folder):
    try:
        rawdata_filenames = [i for i in glob.glob('*')]
        to_check_extension = []
        for file in rawdata_filenames:
            os.rename(file, str.lower(file))
        rawdata_filenames = [i for i in glob.glob('*')]
        # double xls in name
        xls_xls_files = [x for x in rawdata_filenames if ".xls.xls" in x]
        for f_xls_xls in xls_xls_files:
            # print(f_xls)
            os.rename(f_xls_xls, f_xls_xls[:-4])
        rawdata_filenames = [i for i in glob.glob('*')]
        xls_xlsx_files = [x for x in rawdata_filenames if ".xls.xlsx" in x]
        for f_xls_xlsx in xls_xlsx_files:
            # print(f_xls)
            os.rename(f_xls_xlsx, f_xls_xlsx[:-5])
        rawdata_filenames = [i for i in glob.glob('*')]
        xlsx_xls_files = [x for x in rawdata_filenames if ".xlsx.xls" in x]
        for f_xlsx_xls in xlsx_xls_files:
            # print(f_xls)
            os.rename(f_xlsx_xls, f_xlsx_xls[:-4])
        for file in rawdata_filenames:
            if file[-4:] not in [".xls", "xlsx"]:
                to_check_extension.append(file)
            elif file[-4:] == ".xls":
                # rewrite xls to txt
                if outb_actual_file_key in file:
                    base = os.path.splitext(file)[0]
                    os.rename(file, base + '.txt')
            elif file[-4:] == ".xlsx":
                # rewrite xls to txt
                if outb_actual_file_key in file:
                    base = os.path.splitext(file)[0]
                    os.rename(file, base + '.txt')
        rawdata_filenames = [i for i in glob.glob('*')]


        rawdata_filenames = [i for i in glob.glob('*')]
        xls_xlsx_files = [x for x in rawdata_filenames if ".xls" in x]
        xls_files = [x for x in xls_xlsx_files if ".xlsx" not in x]
        for f_xls in xls_files:
            #print(f_xls)
            convert_xls_to_xlsx(f_xls, input_folder)
            os.remove(f_xls)
        rawdata_filenames = [i for i in glob.glob('*')]
    except ValueError:
        messagebox.showinfo("Warning", "Check input files, their names and structure")
    return rawdata_filenames

def format_pack(data, column = 'Arrival mapped'):
    data.loc[data[column].str.contains("pack"), column] = "Co-pack"
    data.loc[data[column].str.contains("Pack"), column] = "Co-pack"


def process_choco_daily__sales(raw_choco_daily__sales,to_load, map_names_choco_pf, DC_filter):
    try:
        choco_daily__sales = raw_choco_daily__sales[raw_choco_daily__sales.iloc[:, 0] == "Total"]
        df = choco_daily__sales
        for col in df.columns:
            if len(df) == df[col].isnull().sum():
                df.drop(col, axis=1, inplace=True)
        choco_daily__sales = df
        choco_daily__sales = choco_daily__sales.rename({"Unnamed: 0": "Product", "Unnamed: 1": "input_Despatch_mapped"},
                                                       axis=1)
        choco_daily__sales["Product"] = "Chocolate"
        choco_daily__sales["Arrival mapped"] = "Customers"
        choco_daily__sales["File name"] = "Outbound Plan"
        # map DC and filter for "Despatch mapped"
        choco_daily__sales = choco_daily__sales.merge(map_names_choco_pf, left_on="input_Despatch_mapped",
                                                      right_on="Choco", how="left")


        choco_daily__sales = choco_daily__sales.rename({"Standard_name": "Despatch mapped"}, axis=1)
        choco_daily__sales.drop(["Choco", "Pet Food", "input_Despatch_mapped"], axis=1, inplace=True)
        choco_daily__sales = choco_daily__sales[choco_daily__sales["Despatch mapped"].isin(DC_filter)]
        choco_daily__sales = pd.melt(choco_daily__sales,
                                     id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                     var_name='Date',
                                     value_name='Supply')
        choco_daily__sales["source"] = to_load[0]

        #format_pack(choco_daily__sales)
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return choco_daily__sales


def process_pf_daily__sales(raw_pf_daily__sales,to_load, map_names_choco_pf, DC_filter):
    try:
        pf_daily__sales = raw_pf_daily__sales[raw_pf_daily__sales.iloc[:, 0] != "Total"]
        pf_daily__sales = pf_daily__sales[~pf_daily__sales.iloc[:, 0].isnull()]

        df = pf_daily__sales
        for col in df.columns:
            if len(df) == df[col].isnull().sum():
                df.drop(col, axis=1, inplace=True)
        pf_daily__sales = df
        pf_daily__sales = pf_daily__sales.rename({"Unnamed: 0": "Product", "Unnamed: 1": "input_Despatch_mapped"},
                                                 axis=1)
        pf_daily__sales["Arrival mapped"] = "Customers"
        pf_daily__sales["File name"] = "Outbound Plan"
        # map DC and filter for "Despatch mapped"
        pf_daily__sales = pf_daily__sales.merge(map_names_choco_pf, left_on="input_Despatch_mapped", right_on="Pet Food", how="left")
        pf_daily__sales = pf_daily__sales.rename({"Standard_name": "Despatch mapped"}, axis=1)
        pf_daily__sales.drop(["Choco", "Pet Food", "input_Despatch_mapped"], axis=1, inplace=True)
        pf_daily__sales = pf_daily__sales[pf_daily__sales["Despatch mapped"].isin(DC_filter)]
        pf_daily__sales = pd.melt(pf_daily__sales,
                                  id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                  var_name='Date',
                                  value_name='Supply')
        pf_daily__sales["source"] = to_load[0]
        #format_pack(pf_daily__sales)
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return pf_daily__sales


def process_choco_daily__ship(raw_choco_daily__ship,to_load, map_names_choco_pf, DC_filter):
    try:
        choco_daily__ship = raw_choco_daily__ship[raw_choco_daily__ship.iloc[:, 1] != "Total"]
        choco_daily__ship = choco_daily__ship[choco_daily__ship.iloc[:, 1] != "Despatch"]
        choco_daily__ship = choco_daily__ship[~choco_daily__ship.iloc[:, 1].isnull()]
        df = choco_daily__ship
        for col in df.columns:
            if len(df) == df[col].isnull().sum():
                df.drop(col, axis=1, inplace=True)
        choco_daily__ship = df
        choco_daily__ship = choco_daily__ship.rename(
            {"Unnamed: 0": "Product", "Unnamed: 1": "input_Despatch_mapped", "Unnamed: 2": "input_Arrival_mapped"},
            axis=1)
        choco_daily__ship["Product"] = "Chocolate"
        choco_daily__ship["File name"] = "Inbound Plan"
        df = choco_daily__ship
        df = df.merge(map_names_choco_pf, left_on="input_Despatch_mapped", right_on="Choco", how="left")
        check_map_df0 = df[["Choco", "Standard_name"]]
        check_map_df0 = check_map_df0[check_map_df0["Standard_name"].isnull()]
        df = df.rename({"Standard_name": "Despatch mapped"}, axis=1)
        df.drop(["Choco", "Pet Food", "input_Despatch_mapped"], axis=1, inplace=True)
        df = df.merge(map_names_choco_pf, left_on="input_Arrival_mapped", right_on="Choco", how="left")
        check_map_df1 = df[["Choco", "Standard_name"]]
        check_map_df1 = check_map_df1[check_map_df1["Standard_name"].isnull()]
        check_map_df = pd.concat([check_map_df0, check_map_df1])
        df = df.rename({"Standard_name": "Arrival mapped"}, axis=1)
        df.drop(["Choco", "Pet Food", "input_Arrival_mapped"], axis=1, inplace=True)
        choco_daily__ship = df

        # map DC and filter for "Despatch mapped"
        choco_daily__ship_outb = choco_daily__ship[choco_daily__ship["Despatch mapped"].isin(DC_filter)]
        choco_daily__ship_outb = pd.melt(choco_daily__ship_outb,
                                         id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                         var_name='Date',
                                         value_name='Supply')
        choco_daily__ship_outb = choco_daily__ship_outb[choco_daily__ship_outb["Arrival mapped"] != "SIT"]
        choco_daily__ship_outb = choco_daily__ship_outb[choco_daily__ship_outb["Despatch mapped"] != "sit"]
        choco_daily__ship_outb["File name"] = "Outbound Plan"
        choco_daily__ship_outb["source"] = to_load[0]
        choco_daily__ship_inb = choco_daily__ship[choco_daily__ship["Arrival mapped"].isin(DC_filter)]
        choco_daily__ship_inb = pd.melt(choco_daily__ship_inb,
                                        id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                        var_name='Date',
                                        value_name='Supply')
        choco_daily__ship_inb = choco_daily__ship_inb[choco_daily__ship_inb["Arrival mapped"] != "SIT"]
        choco_daily__ship_inb = choco_daily__ship_inb[choco_daily__ship_inb["Despatch mapped"] != "sit"]
        choco_daily__ship_inb = choco_daily__ship_inb.rename(
            {"Arrival mapped": "Despatch mappednew", "Despatch mapped": "Arrival mapped"}, axis=1)
        choco_daily__ship_inb = choco_daily__ship_inb.rename(
            {"Despatch mappednew": "Despatch mapped", "Despatch mapped": "Arrival mapped"}, axis=1)
        choco_daily__ship_inb["source"] = to_load[0]
        #format_pack(choco_daily__ship_outb)
        #format_pack(choco_daily__ship_inb)
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return choco_daily__ship_outb, choco_daily__ship_inb, check_map_df




def process_pf_daily__ship(raw_pf_daily__ship,to_load, map_names_choco_pf, DC_filter):
    try:
        pf_daily__ship = raw_pf_daily__ship[raw_pf_daily__ship.iloc[:, 1] != "Total"]
        pf_daily__ship = pf_daily__ship[pf_daily__ship.iloc[:, 1] != "Despatch"]
        pf_daily__ship = pf_daily__ship[~pf_daily__ship.iloc[:, 1].isnull()]
        df = pf_daily__ship
        for col in df.columns:
            if df[col].sum() == 0:
                df.drop(col, axis=1, inplace=True)
        pf_daily__ship = df
        pf_daily__ship = pf_daily__ship.rename(
            {"Unnamed: 0": "Product", "Unnamed: 1": "input_Despatch_mapped", "Unnamed: 2": "input_Arrival_mapped"},
            axis=1)
        pf_daily__ship["File name"] = "Inbound Plan"
        df = pf_daily__ship
        df = df.merge(map_names_choco_pf, left_on="input_Despatch_mapped", right_on="Pet Food", how="left")
        check_map_df0 = df[["Pet Food", "Standard_name"]]
        check_map_df0 = check_map_df0[check_map_df0["Standard_name"].isnull()]
        df = df.rename({"Standard_name": "Despatch mapped"}, axis=1)
        df.drop(["Choco", "Pet Food", "input_Despatch_mapped"], axis=1, inplace=True)
        df = df.merge(map_names_choco_pf, left_on="input_Arrival_mapped", right_on="Pet Food", how="left")
        check_map_df1 = df[["Pet Food", "Standard_name"]]
        check_map_df1 = check_map_df1[check_map_df1["Standard_name"].isnull()]
        check_map_df = pd.concat([check_map_df0, check_map_df1])
        df = df.rename({"Standard_name": "Arrival mapped"}, axis=1)
        df.drop(["Choco", "Pet Food", "input_Arrival_mapped"], axis=1, inplace=True)
        pf_daily__ship = df
        # map DC and filter for "Despatch mapped"
        pf_daily__ship_outb = pf_daily__ship[pf_daily__ship["Despatch mapped"].isin(DC_filter)]
        pf_daily__ship_outb = pd.melt(pf_daily__ship_outb,
                                      id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                      var_name='Date',
                                      value_name='Supply')
        pf_daily__ship_outb["File name"] = "Outbound Plan"
        pf_daily__ship_outb["source"] = to_load[0]
        pf_daily__ship_inb = pf_daily__ship[pf_daily__ship["Arrival mapped"].isin(DC_filter)]
        pf_daily__ship_inb = pd.melt(pf_daily__ship_inb,
                                     id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                                     var_name='Date',
                                     value_name='Supply')
        pf_daily__ship_inb = pf_daily__ship_inb.rename(
            {"Arrival mapped": "Despatch mappednew", "Despatch mapped": "Arrival mapped"}, axis=1)
        pf_daily__ship_inb = pf_daily__ship_inb.rename(
            {"Despatch mappednew": "Despatch mapped", "Despatch mapped": "Arrival mapped"}, axis=1)
        pf_daily__ship_inb["source"] = to_load[0]
        #format_pack(pf_daily__ship_outb)
        #format_pack(pf_daily__ship_inb)
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return pf_daily__ship_outb, pf_daily__ship_inb, check_map_df



def process_sitenka_inb(raw_sitenka_inb, to_load):
    try:
        sitenka_inb = raw_sitenka_inb
        to_fill_date_df = sitenka_inb.iloc[2:3, :]
        sitenka_inb = sitenka_inb[sitenka_inb.iloc[:, 1] != "PRODUCT"]
        for i in range(len(to_fill_date_df.columns) - 1):
            i = i + 1
            if to_fill_date_df.iloc[:, i].isnull().sum() == 1:
                val_to_fill = to_fill_date_df.iloc[:, i - 1]
                to_fill_date_df.iloc[:, i] = val_to_fill
        to_fill_date_list = to_fill_date_df.values.tolist()[0]
        sitenka_inb.columns = to_fill_date_list
        sitenka_inb = sitenka_inb.loc[:, (sitenka_inb.isin(['M', "N", "кол-во траков"]).any())]
        sitenka_inb = sitenka_inb[sitenka_inb.iloc[:, 0] == "кол-во траков"]
        sitenka_inb = sitenka_inb.rename({"PRODUCT": "Product"}, axis=1)
        sitenka_inb["Product"] = "Chocolate"
        sitenka_inb["File name"] = "Inbound Plan"
        sitenka_inb["Arrival mapped"] = "SIT"  # SIT ?
        sitenka_inb["Despatch mapped"] = "SDC"
        sitenka_inb = pd.melt(sitenka_inb,
                              id_vars=["File name", "Product", "Despatch mapped", "Arrival mapped"],
                              var_name='Date',
                              value_name='Truck')
        sitenka_inb["Supply"] = sitenka_inb["Truck"] * 28
        sitenka_inb.drop("Truck", axis=1, inplace=True)
        sitenka_inb["source"] = to_load[0]
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return sitenka_inb




def process_outb_actual(f, map_dc_saps_outb_act, actual_date):
    try:
        outb_actual = pd.read_csv(f, encoding="utf-16", sep='\t')
        outb_actual["source"] = f
        outb_actual = outb_actual[["source", "SAP код склада", "Дата заг (с)", "Кол-во пал", "склад отгрузки"]]
        count_rows = len(outb_actual) - 1
        if np.isnan(outb_actual.iloc[count_rows, 1]):
            outb_actual = outb_actual.iloc[:count_rows, :]
        outb_actual["Дата заг (с)"] = outb_actual["Дата заг (с)"].apply(lambda x: str(x) + "")
        outb_actual["Дата заг (с)"] = outb_actual["Дата заг (с)"].apply(lambda x: parser.parse(x, dayfirst=True))
        outb_actual = outb_actual.rename({"SAP код склада": "SAP", "Дата заг (с)": "Date", "Кол-во пал": "Supply",
                            "склад отгрузки": "input_Despatch_mapped"}, axis=1)
        outb_actual["File name"] = "Outbound Actual"
        outb_actual["Product"] = "All"
        outb_actual = outb_actual.merge(map_dc_saps_outb_act, on="SAP", how="left")
        outb_actual.drop(["SAP", "Input_name"], axis=1, inplace=True)
        outb_actual.rename({"Standard_name": 'Arrival mapped'}, axis=1, inplace=True)
        outb_actual.loc[outb_actual['Arrival mapped'].isnull(), 'Arrival mapped'] = "Customers"  # "Customers Chocolate"
        outb_actual = outb_actual.merge(map_dc_saps_outb_act, left_on="input_Despatch_mapped", right_on="Input_name",
                                        how="left")
        outb_actual.drop(["SAP", "Input_name", "input_Despatch_mapped"], axis=1, inplace=True)
        outb_actual.rename({"Standard_name": 'Despatch mapped'}, axis=1, inplace=True)
        # outb_actual.fillna(0,inplace=True)
        outb_actual = outb_actual.groupby(
            ["source", "File name", "Product", 'Date', "Despatch mapped", "Arrival mapped"]).sum()
        outb_actual.reset_index(inplace=True)
        # filter Date
        outb_actual = outb_actual[outb_actual["Date"] == actual_date]
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + f)
    return outb_actual



def process_inb_actual(f, map_dc_inb_act, actual_date):
    try:
        inb_actual = read_excel_wo_hidden(f, sheet_name=inb_actual_sheet_key, encoding='latin-1')
        inb_actual.index = inb_actual.index + 2
        DC_name = inb_actual.columns[2]
        date_val = inb_actual.columns[4]
        inb_actual = inb_actual[[DC_name, "Unnamed: 3", date_val, "Unnamed: 5"]]
        inb_actual = inb_actual.rename(
            {"Unnamed: 3": "Supply", "Unnamed: 5": "input_Despatch_mapped", date_val: "DOC", DC_name: "DOC1"},
            axis=1)
        inb_actual["Date"] = date_val
        inb_actual["input_Arrival_mapped"] = DC_name
        inb_actual["source"] = f
        # filter data in format yyyy-mm-dd
        inb_actual[inb_actual["Date"] == actual_date]
        inb_actual["File name"] = "Inbound Actual"
        inb_actual["Product"] = "All"
        inb_actual = inb_actual[inb_actual["input_Despatch_mapped"] != "Склад"]
        inb_actual = inb_actual[inb_actual["DOC"] != "пал."]
        inb_actual = inb_actual[inb_actual["DOC1"] != "трак."]


        inb_actual["type_date"] = inb_actual["DOC1"].apply(lambda x: str(type(x)))
        inb_actual_filter_str = inb_actual[inb_actual["type_date"] == "<class 'str'>"]
        inb_actual_filter_str["DOC1"] = inb_actual_filter_str["DOC1"].apply(lambda x: parser.parse(x, dayfirst=True))
        inb_actual_filter_str = inb_actual_filter_str[((inb_actual_filter_str["DOC1"] >= actual_date) & (
                inb_actual_filter_str["DOC1"] < date_offset(actual_date, 1)))]
        inb_actual_filter_nan_time = inb_actual[
            (inb_actual["type_date"] != "<class 'datetime.datetime'>") & (inb_actual["type_date"] != "<class 'str'>")]
        inb_actual_filter_date = inb_actual[(inb_actual["type_date"] == "<class 'datetime.datetime'>")]
        inb_actual_filter_date = inb_actual_filter_date[
            (inb_actual_filter_date["DOC1"] < parser.parse("1900-02-02")) | (
                        (inb_actual_filter_date["DOC1"] >= actual_date) & (
                        inb_actual_filter_date["DOC1"] < date_offset(actual_date, 1)))]
        inb_actual = pd.concat([inb_actual_filter_nan_time, inb_actual_filter_date, inb_actual_filter_str])
        inb_actual = inb_actual.drop(["DOC", "DOC1"], axis=1)
        inb_actual = inb_actual[~(inb_actual["Supply"].isnull())]
        inb_actual["check_null_despatch"] = inb_actual.index - 1
        inb_actual["in_index"] = np.where(inb_actual["check_null_despatch"].isin(inb_actual.index.tolist()), "Y", "N")
        inb_actual["in_index"] = np.where(
            (inb_actual["in_index"] == "N") & (inb_actual["input_Despatch_mapped"].isnull()), "Y", "N")
        to_check_null_despatch_idex_list = inb_actual[inb_actual["in_index"] == "Y"].index.to_list()

        if len(to_check_null_despatch_idex_list) > 0:
            wb = load_workbook(f)
            ws = wb.get_sheet_by_name(inb_actual_sheet_key)
            for col in range(ws.max_column):
                col1 = col + 1
                column_letter = get_column_letter(col1)
                if ws.cell(2, col1).value == "Склад":
                    for val in to_check_null_despatch_idex_list:
                        cell_to_change = ws.cell(val, col1)
                        ws[column_letter+str(val)] ="unmatched"
                        #cell_to_change.value = "unmatched"
                        cell_to_change.font = Font(bold=True, color="FFA07A")
                        cell_to_change.fill = PatternFill("solid", fgColor="800000")
            wb.save(f)
            messagebox.showinfo("Warning", "Please check the next file: " + f + "and rename unmatch storage!")
        inb_actual.loc[to_check_null_despatch_idex_list, "input_Despatch_mapped"] = "unmatched"
        inb_actual = inb_actual.fillna(method='ffill')

        desp_map = map_dc_inb_act[["input_Despatch_mapped", "Standard_name_desp"]]
        inb_actual = inb_actual.merge(desp_map, on="input_Despatch_mapped", how="left")
        check_map_desp_inb_actual = inb_actual[["input_Despatch_mapped", "Standard_name_desp"]]
        check_map_desp_inb_actual = check_map_desp_inb_actual[check_map_desp_inb_actual["Standard_name_desp"].isnull()]
        inb_actual.drop("input_Despatch_mapped", axis=1, inplace=True)
        inb_actual.rename({"Standard_name_desp": 'Despatch mapped'}, axis=1, inplace=True)
        arr_map = map_dc_inb_act[["input_Arrival_mapped", "Standard_name_arr"]]
        inb_actual = inb_actual.merge(arr_map, on="input_Arrival_mapped", how="left")
        check_map_arr_inb_actual = inb_actual[["input_Arrival_mapped", "Standard_name_arr"]]
        check_map_arr_inb_actual = check_map_arr_inb_actual[check_map_arr_inb_actual["Standard_name_arr"].isnull()]
        inb_actual.drop("input_Arrival_mapped", axis=1, inplace=True)
        inb_actual.rename({"Standard_name_arr": 'Arrival mapped'}, axis=1, inplace=True)
        inb_actual = inb_actual.rename({"Arrival mapped": "Despatch mappednew", "Despatch mapped": "Arrival mapped"},
                                       axis=1)
        inb_actual = inb_actual.rename({"Despatch mappednew": "Despatch mapped", "Despatch mapped": "Arrival mapped"},
                                       axis=1)
        inb_actual.loc[inb_actual['Arrival mapped'] == "Co-Packing SDC", 'Arrival mapped'] = "Co-pack"
        inb_actual["Arrival mapped"].fillna("unmatched", inplace=True)
        #format_pack(inb_actual)

    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + f)
    return inb_actual, check_map_arr_inb_actual, check_map_desp_inb_actual


def process_stock(stock_df, to_load):
    try:
        stock_df = stock_df.loc[:, (stock_df.isin(['Warehouse', "Gum and Confections", "Total"]).any())]
        to_fill_name_col_df = stock_df.iloc[0:1, :]
        for i in range(len(to_fill_name_col_df.columns) - 1):
            i = i + 1
            if to_fill_name_col_df.iloc[:, i].isnull().sum() == 1:
                val_to_fill = to_fill_name_col_df.iloc[:, i - 1]
                to_fill_name_col_df.iloc[:, i] = val_to_fill
        to_fill_name_col_list = to_fill_name_col_df.values.tolist()[0]
        stock_df.columns = to_fill_name_col_list
        stock_df = stock_df[stock_df["Warehouse"] != "Warehouse"]
        stock_df = stock_df[~(stock_df["Warehouse"].isnull())]
        stock_df["Start_Stock"] = stock_df["Total"] - stock_df["Gum and Confections"]
        stock_df = stock_df[["Warehouse", "Start_Stock"]]
    except ValueError:
        messagebox.showinfo("Warning", "Please check the next file: " + to_load[0])
    return stock_df


def check_mapping(mapping_file, df_name, sheetname = "inbound_actual", column_input = "input_Despatch_mapped", column_output = "Standard_name_desp"):
    try:
        df_name.drop_duplicates(inplace=True)
        list_name = df_name[column_input].to_list()
        if len(list_name) > 0:
            wb = load_workbook(mapping_file)
            ws = wb.get_sheet_by_name(sheetname)
            for col in range(ws.max_column):
                col1 = col + 1
                column_letter = get_column_letter(col1)
                col_len = len(ws[column_letter])
                for i in range(len(list_name)):
                    if ws.cell(1, col1).value == column_input:
                        ws.cell(col_len + i + 1, col1).value = list_name[i]
                        col_len1 = col_len
                    elif ws.cell(1, col1).value == column_output:
                        cell_to_change = ws.cell(col_len1 + i + 1, col1)
                        cell_to_change.value = "unmatched"
                        cell_to_change.font = Font(bold=True, color = "FFA07A")
                        cell_to_change.fill = PatternFill("solid", fgColor="800000")
            wb.save(mapping_file)
            messagebox.showinfo("Warning", "Please check mapping and rewrite unmatched on the sheet "+sheetname)
    except ValueError:
        messagebox.showinfo("Warning", "Can not check mapping for sheet "+sheetname )






#################################################  EXECUTE MAIN FUNCTION AND INTERFACE  ################################################################################

def proces():
    try:
        actual_date = Entry.get(E1)
        actual_date = parser.parse(actual_date)
        # if use next command, then after each itaration initiated by user variable core_path will be increased
        #core_path = os.getcwd()
        core_path = os.path.dirname(os.path.abspath(__file__))
        #print(core_path)
        folder_with_data = Entry.get(E2)
        # 1. folder with input files (!without \\ at the end!)
        input_folder = core_path + "\\" + folder_with_data
        #print("input_folder^"+input_folder)
        # 2. folder, where the map file is stored (!with \\ at the end!)
        map_folder = core_path + "\\mapping\\"
        # 3. folder, where the created table will be stored (!with \\ at the end!)
        output_folder = core_path + "\\output_of_code\\"
        output_file_name = "WH_data_2019_10_11_stock_v1.xlsx"
        os.chdir(input_folder)
        rawdata_filenames = files_check_and_list(input_folder)
        rawdata_filenames = [i for i in glob.glob('*')]
        #print("1:")
        #print(rawdata_filenames)

        ########### Aggregation of the plan data ########
        DC_filter_df = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="DC_filter", encoding="latin-1")
        DC_filter = DC_filter_df.transpose().values.tolist()[0]
        map_names_choco_pf = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="plan", encoding="latin-1")


        to_load = [k for k in rawdata_filenames if choco_daily_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        #print(to_load,rawdata_filenames,  choco_daily_file_key)
        if len(to_load) == 1:
            raw_choco_daily__sales = pd.read_excel(to_load[0], sheet_name=choco_daily_sales_sheet_key,
                                                   encoding="latin-1")
        choco_daily__sales = process_choco_daily__sales(raw_choco_daily__sales,to_load, map_names_choco_pf, DC_filter)


        to_load = [k for k in rawdata_filenames if pf_daily_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        if len(to_load) == 1:
            raw_pf_daily__sales = pd.read_excel(to_load[0], sheet_name=pf_daily_sales_sheet_key, encoding="latin-1")
        pf_daily__sales = process_pf_daily__sales(raw_pf_daily__sales,to_load, map_names_choco_pf, DC_filter)

        to_load = [k for k in rawdata_filenames if choco_daily_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        if len(to_load) == 1:
            raw_choco_daily__ship = pd.read_excel(to_load[0], sheet_name=choco_daily_ship_sheet_key, encoding="latin-1")
        choco_daily__ship_outb, choco_daily__ship_inb, check_map_plan_choco = process_choco_daily__ship(raw_choco_daily__ship, to_load, map_names_choco_pf, DC_filter)
        check_mapping(map_folder + "Mapping.xlsx", check_map_plan_choco, sheetname="plan",
                      column_input="Choco",
                      column_output="Standard_name")

        to_load = [k for k in rawdata_filenames if pf_daily_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        if len(to_load) == 1:
            raw_pf_daily__ship = pd.read_excel(to_load[0], sheet_name=pf_daily_ship_sheet_key, encoding="latin-1")
        pf_daily__ship_outb, pf_daily__ship_inb, check_map_plan_pf = process_pf_daily__ship(raw_pf_daily__ship,to_load, map_names_choco_pf, DC_filter)
        check_mapping(map_folder + "Mapping.xlsx", check_map_plan_pf, sheetname="plan",
                      column_input="Pet Food",
                      column_output="Standard_name")
        # shift dates due to transit time
        daily__ship_inb = pd.concat([pf_daily__ship_inb, choco_daily__ship_inb])
        transit_df = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="Transit", encoding='latin-1')
        transit_df = transit_df.rename({"FROM": "Arrival mapped", "TO": "Despatch mapped"},
                                                       axis=1)
        daily__ship_inb_2 = daily__ship_inb.merge(transit_df,
                                                  on=["Despatch mapped", "Arrival mapped"],
                                                  how="left")
        # check mapping
        check_transit = daily__ship_inb_2[["Despatch mapped", "Arrival mapped", "offset"]]
        check_transit = check_transit.rename({"Arrival mapped" : "FROM", "Despatch mapped" : "TO"},
                                       axis=1)
        check_transit = check_transit[check_transit["offset"].isnull()]
        check_transit.drop_duplicates(inplace=True)
        list_first_column = check_transit["TO"].to_list()
        list_second_column = check_transit["FROM"].to_list()
        if len(list_first_column) > 0:
            wb = load_workbook(map_folder + "Mapping.xlsx")
            ws = wb.get_sheet_by_name("Transit")
            col_len = len(ws["A"])
            for col in range(ws.max_column):
                col1 = col + 1
                for i in range(len(list_first_column)):
                    if ws.cell(1, col1).value == "TO":
                        ws.cell(col_len + i + 1, col1).value = list_first_column[i]
                    elif ws.cell(1, col1).value == "FROM":
                        cell_to_change = ws.cell(col_len + i + 1, col1)
                        cell_to_change.value = list_second_column[i]
                    elif ws.cell(1, col1).value == "offset":
                        cell_to_change = ws.cell(col_len + i + 1, col1)
                        cell_to_change.value = 0
                        cell_to_change.font = Font(bold=True, color="FFA07A")
                        cell_to_change.fill = PatternFill("solid", fgColor="800000")
            wb.save(map_folder + "Mapping.xlsx")
            messagebox.showinfo("Warning", "Please check mapping and rewrite unmatched on the sheet Transit")

        daily__ship_inb_2["offset"].fillna(0, inplace=True)
        daily__ship_inb_2["Dat"] = daily__ship_inb_2["Date"]
        daily__ship_inb_2["Date"] = daily__ship_inb_2.apply(
            lambda row: date_offset(row["Dat"], row["offset"]) if row["offset"] > 1 else row["Dat"], axis=1)


        to_load = [k for k in rawdata_filenames if operations_SIT_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        if len(to_load) == 1:
            raw_sitenka_inb = pd.read_excel(to_load[0], sheet_name=operations_SIT_sheet_key, encoding="latin-1")
        sitenka_inb = process_sitenka_inb(raw_sitenka_inb, to_load)

        ########## Aggregation of the actual data ##########
        map_dc_saps_outb_act = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="outbound_actual",
                                             encoding='latin-1')
        map_dc_inb_act = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="inbound_actual", encoding='utf-16')
        to_load = [k for k in rawdata_filenames if outb_actual_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        outb_actual = pd.DataFrame()
        for f in to_load:
            raw_outb_actual = process_outb_actual(f, map_dc_saps_outb_act, actual_date)
            outb_actual = pd.concat([outb_actual, raw_outb_actual])


        to_load = [k for k in rawdata_filenames if inb_actual_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        inb_actual = pd.DataFrame()
        check_map_arr_inb_actual_all = pd.DataFrame()
        check_map_desp_inb_actual_all = pd.DataFrame()
        for f in to_load:
            raw_inb_actual, check_map_arr_inb_actual, check_map_desp_inb_actual = process_inb_actual(f, map_dc_inb_act, actual_date)
            inb_actual = pd.concat([inb_actual, raw_inb_actual])
            check_map_arr_inb_actual_all = pd.concat([check_map_arr_inb_actual_all, check_map_arr_inb_actual])
            check_map_desp_inb_actual_all = pd.concat([check_map_desp_inb_actual_all, check_map_desp_inb_actual])

        check_mapping(map_folder + "Mapping.xlsx", check_map_arr_inb_actual_all, sheetname="inbound_actual",
                      column_input="input_Arrival_mapped",
                      column_output="Standard_name_arr")
        check_mapping(map_folder + "Mapping.xlsx", check_map_desp_inb_actual_all, sheetname="inbound_actual",
                      column_input="input_Despatch_mapped",
                      column_output="Standard_name_desp")






        to_load = [k for k in rawdata_filenames if stock_monitor_file_key in k]
        to_load = [k for k in to_load if "~" not in k]
        if len(to_load) == 1:
            stock_df = pd.read_excel(to_load[0], encoding="latin-1")
        stock_df = process_stock(stock_df, to_load)

        ################ Merge all data in one file ######################
        all_df = pd.concat([daily__ship_inb_2,
                            sitenka_inb,
                            pf_daily__ship_outb,
                            choco_daily__ship_outb,
                            pf_daily__sales,
                            choco_daily__sales,
                            inb_actual,
                            outb_actual])
        all_df.loc[all_df['Product'] == "Dry", 'Product'] = "Pet Food Dry"
        all_df.loc[all_df['Product'] == "Wet", 'Product'] = "Pet Food Wet"
        all_df.loc[all_df['Product'] == "Import", 'Product'] = "Pet Food Import"
        all_df = all_df[['File name', 'Product', 'Date', 'Despatch mapped', 'Arrival mapped', 'Supply']]
        all_df = all_df.merge(DC_filter_df, how="left", on="Despatch mapped")
        new = all_df["File name"].str.split(" ", n=1, expand=True)
        all_df["Plan|Act"] = new[1]
        all_df["Inb|Outb"] = new[0]
        all_df["Supply"] = np.where(all_df["Supply"] == '', np.nan, all_df["Supply"])
        all_df["Supply"] = np.where(all_df["Supply"] == ' ', np.nan, all_df["Supply"])
        all_df["Supply"] = np.where(all_df["Supply"] == '  ', np.nan, all_df["Supply"])
        all_df["Supply"] = all_df["Supply"].astype(float)
        df_to_calc_change_stock = all_df[['File name', 'Date', 'Despatch mapped', 'Supply']]
        df_to_calc_change_stock = all_df[['File name', 'Date', 'Despatch mapped', 'Supply']]
        df_to_calc_change_stock = df_to_calc_change_stock.groupby(['File name', 'Date', 'Despatch mapped']).sum()
        df_to_calc_change_stock.reset_index(inplace=True)
        df_to_calc_change_stock = df_to_calc_change_stock.rename({'Supply': "change_stock"}, axis=1)
        new = df_to_calc_change_stock["File name"].str.split(" ", n=1, expand=True)
        df_to_calc_change_stock["In|Out"] = new[0]
        df_to_calc_change_stock["Plan|Act"] = new[1]
        df_to_calc_change_stock = df_to_calc_change_stock[
            ~((df_to_calc_change_stock["Plan|Act"] == "Plan") & (df_to_calc_change_stock["Date"] == actual_date))]
        df_to_calc_change_stock.loc[df_to_calc_change_stock["In|Out"] == "Outbound", "change_stock"] = \
        df_to_calc_change_stock["change_stock"] * (-1)
        df_to_calc_change_stock = df_to_calc_change_stock[['Plan|Act', 'Date', 'Despatch mapped', "change_stock"]]
        df_to_calc_change_stock = df_to_calc_change_stock.groupby(['Plan|Act', 'Date', 'Despatch mapped']).sum()
        df_to_calc_change_stock.reset_index(inplace=True)
        df_to_calc_change_stock["Date"] = df_to_calc_change_stock["Date"].apply([lambda x: date_offset(x, 1)])
        df_to_calc_change_stock["change_stock"] = np.where(df_to_calc_change_stock["Date"] == actual_date, 0,
                                                           df_to_calc_change_stock["change_stock"])
        change_stock_greater_than_actual_date = df_to_calc_change_stock[df_to_calc_change_stock["Date"] >= actual_date]
        change_stock_greater_than_actual_date.reset_index(inplace=True, drop=True)
        df = change_stock_greater_than_actual_date
        df["cum_change_stock"] = 0
        for i in range(len(df)):
            #pl_ac_val = df.loc[i, "Plan|Act"]
            date_val = df.loc[i, "Date"]
            desp_val = df.loc[i, "Despatch mapped"]
            df_0 = df[(df["Date"] <= date_val) & (df["Despatch mapped"] == desp_val)]
            df_0 = df_0[['Despatch mapped', "change_stock"]]
            df_0 = df_0.groupby(['Despatch mapped']).sum()
            cum_change_stock = df_0.iloc[0, 0]
            df.loc[i, "cum_change_stock"] = cum_change_stock
        final_change_stock = df
        final_change_stock = final_change_stock[['Date', 'Despatch mapped', 'cum_change_stock']]
        final_change_stock = final_change_stock[final_change_stock["cum_change_stock"] != 0]

        df_to_calc_Throughput = all_df[['File name', 'Date', 'Despatch mapped', 'Supply']]
        new = df_to_calc_Throughput["File name"].str.split(" ", n=1, expand=True)
        df_to_calc_Throughput["Plan|Act"] = new[1]
        df_to_calc_Throughput = df_to_calc_Throughput[["Plan|Act", 'Date', 'Despatch mapped', 'Supply']]
        df_to_calc_Throughput = df_to_calc_Throughput.groupby(["Plan|Act", 'Date', 'Despatch mapped']).sum()
        df_to_calc_Throughput.reset_index(inplace=True)
        df_to_calc_Throughput = df_to_calc_Throughput.rename({'Supply': "Throughput"}, axis=1)
        all_df = all_df.merge(df_to_calc_Throughput, how="left", on=["Plan|Act", 'Date', 'Despatch mapped'])
        all_df = all_df.merge(final_change_stock, how="left", on=['Date', 'Despatch mapped'])
        all_df = all_df.merge(stock_df, how="left", on=['Warehouse'])
        all_df["cum_change_stock"].fillna(0, inplace=True)
        all_df["Stock"] = all_df["cum_change_stock"] + all_df["Start_Stock"]

        all_df = all_df[['File name', 'Product', 'Date', 'Despatch mapped', 'Arrival mapped',
                         'Supply', 'Throughput', 'Total Space', 'Stock']]
        all_df["Free Space"] = all_df["Total Space"] - all_df["Stock"]
        all_df["SOC Utilization"] = all_df["Stock"] / all_df["Total Space"]
        all_df.loc[all_df['Arrival mapped'] == "Customers", 'Arrival mapped'] = "Customers Pet Nutrition"
        all_df.loc[(all_df['Arrival mapped'] == "Customers Pet Nutrition") & (
                    all_df['Product'] == "Chocolate"), 'Arrival mapped'] = "Customers Chocolate"
        all_df.loc[(all_df['Arrival mapped'] == "Customers Pet Nutrition") & (
                    all_df['Product'] == "All"), 'Arrival mapped'] = "Customers Chocolate"

        format_pack(all_df, column='Despatch mapped')
        format_pack(all_df, column='Arrival mapped')
        # rename DCs to full
        map_to_report = pd.read_excel(map_folder + "Mapping.xlsx", sheet_name="map_to_report", encoding="latin-1")
        map_to_report_arr = map_to_report[["Arrival mapped", "Arrival mapped in report"]]
        all_df = all_df.merge(map_to_report_arr, how="left", on="Arrival mapped")
        check_map_to_report_arr = all_df[["Arrival mapped", "Arrival mapped in report"]]
        check_map_to_report_arr = check_map_to_report_arr[check_map_to_report_arr["Arrival mapped in report"].isnull()]
        check_mapping(map_folder + "Mapping.xlsx", check_map_to_report_arr, sheetname="map_to_report",
                      column_input="Arrival mapped",
                      column_output="Arrival mapped in report")
        all_df.drop("Arrival mapped", axis=1, inplace=True)
        all_df.rename(columns={"Arrival mapped in report": "Arrival mapped"}, inplace=True)
        all_df["Arrival mapped"].fillna("unmatched", inplace = True)
        map_to_report_des = map_to_report[["Arrival mapped", "Arrival mapped in report"]]
        map_to_report_des = map_to_report_des[map_to_report_des["Arrival mapped"].isin(DC_filter)]
        map_to_report_des.rename(
            columns={"Arrival mapped": "Despatch mapped", "Arrival mapped in report": "Despatch mapped in report"},
            inplace=True)
        all_df = all_df.merge(map_to_report_des, how="left", on="Despatch mapped")
        check_map_to_report_des = all_df[["Despatch mapped", "Despatch mapped in report"]]
        check_map_to_report_des = check_map_to_report_des[check_map_to_report_des["Despatch mapped in report"].isnull()]
        check_map_to_report_des.rename(
            columns={"Despatch mapped":"Arrival mapped", "Despatch mapped in report":"Arrival mapped in report"},
            inplace=True)
        check_mapping(map_folder + "Mapping.xlsx", check_map_to_report_des, sheetname="map_to_report",
                      column_input="Arrival mapped",
                      column_output="Arrival mapped in report")


        all_df.drop("Despatch mapped", axis=1, inplace=True)
        all_df.rename(columns={"Despatch mapped in report": "Despatch mapped"}, inplace=True)
        all_df["Despatch mapped"].fillna("unmatched", inplace=True)
        all_df.fillna(0, inplace=True)
        all_df = all_df[['File name', 'Product', 'Date', 'Despatch mapped', 'Arrival mapped',
                         'Supply', 'Throughput', 'Total Space', 'Stock', 'Free Space', 'SOC Utilization']]

        DC_filter = map_to_report_des.transpose().values.tolist()[1]
        DC_sort_list = DC_filter
        report_all_df = all_df[['File name', 'Product', 'Date', 'Despatch mapped', 'Arrival mapped',
                                'Supply', 'Throughput', 'Total Space', 'Stock', 'Free Space',
                                'SOC Utilization']]
        new = report_all_df["File name"].str.split(" ", n=1, expand=True)
        report_all_df["In|Out"] = new[0]
        report_all_df["Plan|Act"] = new[1]
        # filter right dates
        # inb/outb actual - actual_date
        # inb/outb plan - greater than actual_date
        report_all_df = report_all_df[
            ((report_all_df['Plan|Act'] == "Actual") & (report_all_df['Date'] == actual_date)) |
            ((report_all_df['Plan|Act'] == "Plan") & (report_all_df['Date'] > actual_date))]
        uniq_cols_df = report_all_df[['Date', 'Despatch mapped',
                                      'Throughput', 'Total Space', 'Stock', 'Free Space',
                                      'SOC Utilization']]
        uniq_cols_df.drop_duplicates(inplace=True)
        uniq_cols_df.drop("Total Space", axis=1, inplace=True)
        uniq_cols_df_melt = pd.melt(uniq_cols_df,
                                    id_vars=["Date", "Despatch mapped"],
                                    var_name='Arrival mapped',
                                    value_name='Values')
        uniq_cols_df_melt = uniq_cols_df_melt.set_index(["Despatch mapped", 'Arrival mapped'])
        uniq_cols_df_to_report = multiindex_pivot(uniq_cols_df_melt, columns="Date", values="Values")
        uniq_cols_df_to_report.reset_index(inplace=True)
        cols_to_agg_df = report_all_df[['In|Out', 'Date', 'Despatch mapped', 'Arrival mapped', 'Supply']]
        agg_df = cols_to_agg_df.groupby(['In|Out', 'Date',
                                         'Despatch mapped', 'Arrival mapped']).sum()
        agg_df.reset_index(inplace=True)
        agg_df.set_index(['Despatch mapped', 'In|Out', 'Arrival mapped'], inplace=True)
        fin_report_df = multiindex_pivot(agg_df, columns="Date", values="Supply")
        fin_report_df.reset_index(inplace=True)
        fin_report_df_in = fin_report_df[fin_report_df["In|Out"] == "Inbound"]
        fin_report_df_in.drop("In|Out", axis=1, inplace=True)
        fin_report_df_out = fin_report_df[fin_report_df["In|Out"] == "Outbound"]
        fin_report_df_out.drop("In|Out", axis=1, inplace=True)

        structure_df = fin_report_df[['In|Out', 'Arrival mapped']]
        structure_df = structure_df.drop_duplicates()
        structure_in = structure_df.loc[structure_df['In|Out'] == "Inbound", ['Arrival mapped']]
        structure_out = structure_df.loc[structure_df['In|Out'] == "Outbound", ['Arrival mapped']]
        structure_in["Arrival mapped"] = np.where(structure_in["Arrival mapped"] == 0, "unmatched",
                                                  structure_in["Arrival mapped"])
        structure_out["Arrival mapped"] = np.where(structure_out["Arrival mapped"] == 0, "unmatched",
                                                   structure_out["Arrival mapped"])
        structure_in["Arrival mapped"] = structure_in["Arrival mapped"].astype(object)
        structure_out["Arrival mapped"] = structure_out["Arrival mapped"].astype(object)

        structure_in.sort_values("Arrival mapped", inplace=True)
        structure_in = structure_in[~(structure_in["Arrival mapped"].isin(DC_sort_list))]
        DC_sort_list_df = pd.DataFrame({"Arrival mapped":DC_sort_list})
        structure_in = pd.concat([DC_sort_list_df, structure_in])

        start_structure_out_list = ["Customers Chocolate", "Customers Pet Nutrition"] + DC_sort_list
        start_structure_out_df = pd.DataFrame({"Arrival mapped": start_structure_out_list})
        structure_out.sort_values("Arrival mapped", inplace=True)
        structure_out = structure_out[~(structure_out["Arrival mapped"].isin(start_structure_out_list))]
        structure_out = pd.concat([start_structure_out_df, structure_out])

        df_to_calc_in_out_total = all_df[['File name', 'Date', 'Despatch mapped', 'Supply']]
        df_to_calc_in_out_total = df_to_calc_in_out_total.groupby(['File name', 'Date', 'Despatch mapped']).sum()
        df_to_calc_in_out_total.reset_index(inplace=True)
        df_to_calc_in_out_total = df_to_calc_in_out_total.rename({'Supply': "sum"}, axis=1)
        new = df_to_calc_in_out_total["File name"].str.split(" ", n=1, expand=True)
        df_to_calc_in_out_total["In|Out"] = new[0]
        df_to_calc_in_out_total["Plan|Act"] = new[1]
        df_to_calc_in_out_total = df_to_calc_in_out_total[
            ~((df_to_calc_in_out_total["Plan|Act"] == "Plan") & (df_to_calc_in_out_total["Date"] == actual_date))]
        df_to_calc_in_out_total = df_to_calc_in_out_total[['In|Out', 'Date', 'Despatch mapped', "sum"]]
        df_to_calc_in_out_total = df_to_calc_in_out_total.groupby(['In|Out', 'Date', 'Despatch mapped']).sum()
        df_to_calc_in_out_total.reset_index(inplace=True)
        df_to_calc_in_out_total.set_index(['In|Out', 'Despatch mapped'], inplace=True)
        df_to_calc_in_out_total = df_to_calc_in_out_total[df_to_calc_in_out_total["Date"] >= actual_date]
        in_out_total_df = multiindex_pivot(df_to_calc_in_out_total, columns="Date", values="sum")
        in_out_total_df.reset_index(inplace=True)
        in_out_total_df.rename(columns={'Despatch mapped': 'Arrival mapped'}, inplace=True)
        in_total_df = in_out_total_df[in_out_total_df["In|Out"] == "Inbound"]
        out_total_df = in_out_total_df[in_out_total_df["In|Out"] == "Outbound"]
        in_total_df.drop("In|Out", axis=1, inplace=True)
        out_total_df.drop("In|Out", axis=1, inplace=True)
        DC_filter_df = DC_filter_df.merge(map_to_report_des, how="left", on="Despatch mapped")
        DC_filter_df.drop("Despatch mapped", axis=1, inplace=True)
        DC_filter_df.rename(columns={"Despatch mapped in report": "Despatch mapped"}, inplace=True)


        #################### CREATE REPORT #######################
        merge_4_report = pd.DataFrame()
        for dc in DC_sort_list:
            # inbound
            fin_report_df_in_1_dc = fin_report_df_in.loc[fin_report_df_in["Despatch mapped"] == dc, :]
            in_df_1_dc = structure_in.merge(fin_report_df_in_1_dc, on='Arrival mapped', how="left")
            in_df_1_dc.drop("Despatch mapped", axis=1, inplace=True)
            total_space_val = DC_filter_df.loc[DC_filter_df["Despatch mapped"] == dc, "Total Space"].to_list()[0]
            to_add_new_rows_start = pd.DataFrame({'Arrival mapped': ["", dc, total_space_val, "INBOUND"]})
            in_total_df_1_dc = in_total_df.loc[in_total_df['Arrival mapped'] == dc, :]
            in_total_df_1_dc['Arrival mapped'] = "total"

            to_add_new_rows_end = pd.DataFrame({'Arrival mapped': ["OUTBOUND"]})
            in_df_1_dc = pd.concat([to_add_new_rows_start, in_df_1_dc, in_total_df_1_dc, to_add_new_rows_end])

            # outbound
            fin_report_df_out_1_dc = fin_report_df_out.loc[fin_report_df_out["Despatch mapped"] == dc, :]
            out_df_1_dc = structure_out.merge(fin_report_df_out_1_dc, on='Arrival mapped', how="left")
            out_df_1_dc.drop("Despatch mapped", axis=1, inplace=True)
            to_add_new_rows_end = pd.DataFrame({'Arrival mapped': [""]})
            out_total_df_1_dc = out_total_df.loc[out_total_df['Arrival mapped'] == dc, :]
            out_total_df_1_dc['Arrival mapped'] = "total"
            out_df_1_dc = pd.concat([out_df_1_dc, out_total_df_1_dc, to_add_new_rows_end])
            in_out_df_1_dc = pd.concat([in_df_1_dc, out_df_1_dc])

            # add stock, SOC, free space ..
            uniq_cols_df_to_report_1_dc = uniq_cols_df_to_report.loc[uniq_cols_df_to_report["Despatch mapped"] == dc, :]
            uniq_cols_df_to_report_1_dc.drop("Despatch mapped", axis=1, inplace=True)
            # "Throughput", "Stock", "Free space", "SOC utilization"
            Throughput_1_dc = uniq_cols_df_to_report_1_dc[uniq_cols_df_to_report_1_dc["Arrival mapped"] == "Throughput"]
            Stock_1_dc = uniq_cols_df_to_report_1_dc[uniq_cols_df_to_report_1_dc["Arrival mapped"] == "Stock"]
            Free_Space_1_dc = uniq_cols_df_to_report_1_dc[uniq_cols_df_to_report_1_dc["Arrival mapped"] == "Free Space"]
            SOC_Utilization_1_dc = uniq_cols_df_to_report_1_dc[
                uniq_cols_df_to_report_1_dc["Arrival mapped"] == "SOC Utilization"]
            # uniq_cols_df_to_report_1_dc.rename(columns = {"DM":"Despatch mapped"},inplace=True)
            in_out_df_1_dc = pd.concat(
                [in_out_df_1_dc, Throughput_1_dc, Stock_1_dc, Free_Space_1_dc, SOC_Utilization_1_dc])
            merge_4_report = pd.concat([merge_4_report, in_out_df_1_dc])

        merge_4_report.reset_index(inplace=True, drop=True)
        merge_4_report.index = merge_4_report.index + 2
        #################### FORMAT REPORT #######################
        Soc_st = NamedStyle(name="Soc_st")
        Soc_st.font = Font(bold=True, size=9)
        Soc_st.fill = PatternFill("solid", fgColor="C0C0C0")
        Soc_st.number_format = '0%'

        def dc_style(color_val, i):
            st_mask = NamedStyle(name="st_mask" + str(i))
            st_mask.font = Font(bold=True, size=9)
            st_mask.fill = PatternFill("solid", fgColor=color_val)
            return st_mask

        short_date_st = NamedStyle(name="short_date_st")
        short_date_st.font = Font(size=9, color="ffffff")
        short_date_st.fill = PatternFill("solid", fgColor="008000")
        short_date_st.number_format = '%a'

        wb_ws_name = str(actual_date.day) + "." + str(actual_date.month) + " SOC utilization report"
        merge_4_report.to_excel(output_folder + wb_ws_name + ".xlsx",
                                sheet_name=wb_ws_name,
                                index=False)
        wb = load_workbook(output_folder + wb_ws_name + ".xlsx")
        ws = wb.get_sheet_by_name(wb_ws_name)
        for i in range(ws.max_row + 100):
            i = i + 1
            row = ws.row_dimensions[i]
            row.font = Font(size=9)
            row.number_format = '# ##0'

        for i in range(1, ws.max_row + 1):
            for j in range(1, ws.max_column + 1):
                ws.cell(row=i, column=j).font = Font(size=9)
                ws.cell(row=i, column=j).number_format = '# ##0'
        ws["A1"] = ""

        list_rows = merge_4_report[merge_4_report["Arrival mapped"] == "total"].index.to_list()
        for row_num in list_rows:
            for j in range(1, ws.max_column + 1):
                total_row_st = ws.cell(row=row_num, column=j)
                total_row_st.font = Font(bold=True, size=9)
                bd = Side(border_style='thin', color="000000")
                total_row_st.border = Border(top=bd, bottom=bd)
                total_row_st.number_format = '# ##0'

        list_rows = merge_4_report[merge_4_report["Arrival mapped"] == "Throughput"].index.to_list()
        for row_num in list_rows:
            for j in range(1, ws.max_column + 1):
                Throughput_st = ws.cell(row=row_num, column=j)
                Throughput_st.font = Font(bold=True, size=9)
                Throughput_st.fill = PatternFill("solid", fgColor="FFFF99")
                bd = Side(border_style='thin', color="000000")
                Throughput_st.border = Border(top=bd, bottom=bd)
                Throughput_st.number_format = '# ##0'
        list_rows2 = []
        for num in list_rows:
            list_rows2.append(num - 1)

        for row_num in list_rows2:
            for j in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=j).style = Soc_st

        list_rows = merge_4_report[merge_4_report["Arrival mapped"] == "Stock"].index.to_list()
        for row_num in list_rows:
            for j in range(1, ws.max_column + 1):
                Stock_st = ws.cell(row=row_num, column=j)
                Stock_st.font = Font(bold=True, size=9, color="0070c0")
                Stock_st.fill = PatternFill("solid", fgColor="FFCC99")
                Stock_st.number_format = '# ##0'

        list_rows = merge_4_report[merge_4_report["Arrival mapped"] == "Free Space"].index.to_list()
        for row_num in list_rows:
            for j in range(1, ws.max_column + 1):
                Free_Space_st = ws.cell(row=row_num, column=j)
                Free_Space_st.font = Font(bold=True, size=9, color="375623")
                Free_Space_st.fill = PatternFill("solid", fgColor="FFCC99")
                Free_Space_st.number_format = '# ##0'

        list_rows = merge_4_report[(merge_4_report["Arrival mapped"] == "SOC Utilization")].index.to_list()
        for row_num in list_rows:
            for j in range(1, ws.max_column + 1):
                ws.cell(row=row_num, column=j).style = Soc_st

        dc_colors_list = ["99CCFF", "FF99CC", "99CC00", "FFFF00", "e9d3b4", "99CCFF", "FF99CC", "99CC00", "FFFF00",
                          "e9d3b4", "99CCFF", "FF99CC", "99CC00", "FFFF00", "e9d3b4", "99CCFF", "FF99CC", "99CC00",
                          "FFFF00", "e9d3b4"]
        list_rows = merge_4_report[(merge_4_report["Arrival mapped"] == "INBOUND") | (
                    merge_4_report["Arrival mapped"] == "OUTBOUND")].index.to_list()
        for i in range(len(DC_filter)):
            list_rows_1 = list_rows[i * 2:i * 2 + 2]
            dc_name_index = min(list_rows_1) - 2
            total_space_index = min(list_rows_1) - 1
            dc_style_i = dc_style(dc_colors_list[i], i)
            for row_num in list_rows_1:
                for j in range(1, ws.max_column + 1):
                    ws.cell(row=row_num, column=j).style = dc_style_i
            ws.cell(row=dc_name_index, column=1).style = dc_style_i
            for j in range(2, ws.max_column + 1):
                full_date_st = ws.cell(row=dc_name_index, column=j)
                full_date_st.value = ws.cell(row=1, column=j).value
                # full_date_st.value = full_date_st.value.strftime('%d-%b')
                full_date_st.font = Font(size=9, color="ffffff")
                full_date_st.fill = PatternFill("solid", fgColor="008000")
                full_date_st.number_format = 'dd-mmm'
                short_date_st = ws.cell(row=dc_name_index + 1, column=j)
                short_date_st.value = ws.cell(row=1, column=j).value
                short_date_st.font = Font(size=9, color="ffffff")
                short_date_st.fill = PatternFill("solid", fgColor="008000")
                short_date_st.value = str(short_date_st.value.strftime('%a'))
                short_date_st.alignment = Alignment(horizontal='right')
            Total_Space_st = ws.cell(row=total_space_index, column=1)
            Total_Space_st.font = Font(bold=True, size=9, color="0000FF")
            Total_Space_st.number_format = '# ##0'

        for j in range(1, ws.max_column + 1):
            ws.cell(row=1, column=j).value = ""
            ws.cell(row=1, column=j).style = Soc_st
            ws.cell(row=2, column=j).style = Soc_st

        ws.column_dimensions["A"].width = 30

        wb.save(output_folder + wb_ws_name + ".xlsx")
        messagebox.showinfo("SUCCESS", "Report is created!")

    except ValueError:
        messagebox.showinfo("Warning","Please enter the value in an appropriate format")
top = Tk()
top.title("AGGREGATOR")
#L1 = Label(top, text="My calculator",).grid(row=0,column=1)
L2 = Label(top, text="Actual Date",).grid(row=1,column=0)
L3 = Label(top, text="Folder Name",).grid(row=2,column=0)
#L4 = Label(top, text="Operator",).grid(row=3,column=0)
#L4 = Label(top, text="Answer",).grid(row=4,column=0)
E1 = Entry(top, bd =5)
E1.grid(row=1,column=1)
E2 = Entry(top, bd =5)
E2.grid(row=2,column=1)
#E3 = Entry(top, bd =5)
#E3.grid(row=3,column=1)
#E4 = Entry(top, bd =5)
#E4.grid(row=4,column=1)
B = Button(top, text ="Create Report",command = proces).grid(row=5,column=1,)
top.mainloop()




